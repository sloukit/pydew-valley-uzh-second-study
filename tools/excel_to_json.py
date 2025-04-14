import json
import os
import sys

from impl_excel_to_json.cell_parser import parse_cell
from impl_excel_to_json.excel_helpers import cell_value
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FIRST_DATA_ROW = 3  # in which row data begins
FIRST_DATA_COL = 3  # in which column data begins (numeric)
KEY_COL = "B"  # which column refers to the keys (alphabetical)

# Debugging: False=collects errors; True=aborts on first with stacktrace
EAGER_FAIL = False

if len(sys.argv) > 1:
    excel_filepath = sys.argv[1]
    print(f"Trying to convert: '{excel_filepath}'")
else:
    print("Usage: python excel-convert.py <file>")
    print("")
    print("   Example: python excel_to_json.py 'data/game design.xlsx'")
    exit(1)

# Load the existing Excel file
# workbook = load_workbook("data/game levels.xlsx")
try:
    workbook = load_workbook(excel_filepath)
except Exception as e:
    print("\nERROR: could not open Excel file. Reason:")
    print(f"\t{e}")
    exit(1)

# print(workbook.sheetnames)
# print(workbook.active)
# sheet = workbook.active  # active tab in Excel
final = []
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    print("------------------------------------")
    print(f"Processing sheet: '{sheet}'")

    headers = [
        parse_cell("level_name_text", cell_value(sheet, cell.coordinate))
        for cell in sheet["2"]
        if cell.column >= FIRST_DATA_COL  # Start from column B (column index >= 2)
    ]

    print(f"headers: {headers}")

    col_letters = [
        get_column_letter(i) for i in range(FIRST_DATA_COL, sheet.max_column + 1)
    ]

    levels = [{"level_name_text": h} for h in headers]
    errors = []

    for row in range(FIRST_DATA_ROW, sheet.max_row):
        key = sheet[f"{KEY_COL}{row}"].value

        # when there is no data in `B` column, assume that this is the end of data range and skip to next sheet
        # sometimes `sheet.max_row` is not detect correctly
        if not key:
            break

        for i, col in enumerate(col_letters):
            address = f"{col}{row}"
            cell = cell_value(sheet, address)

            if EAGER_FAIL:
                val = parse_cell(key, cell)
                levels[i][key] = val
            else:
                try:
                    val = parse_cell(key, cell)
                    levels[i][key] = val
                except ValueError as e:
                    errors.append(f"Cell {address}: {e}")
                except Exception as e:
                    errors.append(f"Cell {address}: unexpected error - {e}")

    # Use the commented-out part if the JSON output should be nested document.
    final.append(levels)

    if errors:
        print("\nERROR: failed to validate input format:")
        for e in errors:
            print(f"* {e}")
        exit(1)

print("------------------------------------")
print("Success!")

# print(json.dumps(final, indent=4))

# Directory and file paths
directory = "tools/output"
file_path = os.path.join(directory, "rounds_config.json")
os.makedirs(directory, exist_ok=True)

# Write JSON to the file
with open(file_path, "w") as file:
    json.dump(final, file, indent=4)  # Writing with pretty-printing

print(f"JSON data written to {file_path}")
